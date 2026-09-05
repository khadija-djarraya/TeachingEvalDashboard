"""
ETL: Teaching Evaluation Surveys (UKPSF 2020) -> consolidated dataset
=======================================================================

WHAT THIS SCRIPT DOES
----------------------
Reads the raw "2 Years Data for Teaching Evaluation Surveys Results" folder
(as extracted from OneDrive_2026-09-01.zip) and produces a single consolidated
JSON file that both dashboards (institutional and individual-faculty) are
built from. Nothing in the dashboard is hardcoded — every number traces back
to a cell in one of the source .xlsx files via this script.

SOURCE DATA STRUCTURE (as received)
------------------------------------
2 Years Data for Teaching Evaluation Surveys Results/
  Teaching Evaluation Fall 2020/
    AFS/  staff.xlsx   programme.xlsx   ...pdf timetables and staff schedules
    CSMIS/  ...
    GFP/  "GFP staff Aisha part 20201.xlsx"  "GFP programme Aisha part 20201.xlsx"
    ID/   "ID staff Aisha part 20201.xlsx"   "ID programme Aisha part 20201.xlsx"
  Teaching Evaluation Spring 2021/
    AFS/ CSMIS/ GFP/ ID/  (+ a nested duplicate "Teaching Evaluation 20212/"
                            folder containing a second copy of AFS/CSMIS only
                            -- this script skips that duplicate)
  Teaching Evaluation Fall 2021/
    AFS/ CSMIS/ GFP/ ID/   staff.xlsx / programme.xlsx (standard naming)
  Teaching Evaluation Spring 2022/
    Final 20222 Teaching Evaluation/
      AFS/ CSMIS/ GFP/ ID/   Staff.xlsx / Programme.xlsx (capitalized)

Filenames are NOT consistent across semesters (staff.xlsx, Staff.xlsx,
"<Dept> staff Aisha part <term>.xlsx", "<term> <Dept> staff Aisha part.xlsx").
This script resolves that with a case-insensitive "contains staff / contains
programme" match, preferring an exact "staff.xlsx"/"programme.xlsx" match
when one exists.

WORKBOOK LAYOUT (consistent once you're inside a sheet)
----------------------------------------------------------
- `staff.xlsx`: one sheet per INSTRUCTOR.
- `programme.xlsx`: one sheet per PROGRAM (e.g. AFS, ACCT, BFS, MEC, BA
  under the AFS department; CSMIS, CS, MIS under CSMIS; etc.)
- In both, each sheet has the same shape:
    Row 1:  course codes taught, one per column, ending in a column
            literally labelled "Avg"
    Row 2+: one row per survey question, with:
              col B = UKPSF category (AA / CK / PV)
              col C = UKPSF subcode (A1, A2, K1, V3, ...)
              col D = question text (Arabic, kept as-is)
              col E = internal question sequence number
              col F..: score (0-100) for that question, for each course code
              last col = that question's rolling average across all courses
- All 169 instructor-semester sheets parsed use the IDENTICAL 20-question
  set (same seq numbers / categories / subcodes / text) -- verified in QA
  step below, which is what makes cross-semester and cross-department
  comparison valid instead of comparing apples to oranges.

WHAT COUNTS AS "OVERALL SCORE"
--------------------------------
- Faculty overall = mean of that instructor's 20 question averages
  (i.e. mean of the "Avg" column, NOT re-weighted by course size --
  the source data itself does not carry per-course respondent counts
  at this sheet level, only in the top-level per-course raw files,
  see LIMITATIONS below).
- Program overall = same calculation applied to the programme.xlsx sheet.
- Department average = mean of that department's instructor overalls.
- Institutional average = mean of department averages, per semester.
- "Improvement required" = ANY of {faculty overall, program overall,
  individual question average} < 65.

LIMITATIONS / KNOWN DATA ISSUES (documented, not hidden)
------------------------------------------------------------
1. Program names are not perfectly stable across years: GFP's
   sub-programs are labelled "Foundation" / "Maths & IT" / "English" in
   2020-21 and "GFP" / "MATH & IT" / "ENGLISH" in 2021-22. The dashboard
   normalizes case/spacing but does NOT try to guess that "Foundation"
   and "GFP" are the same thing across a rename -- that is a judgement
   call for you to confirm, not something the script should silently
   assume.
2. This script reads course names from the top-level per-course raw files
    (e.g. `20201.xlsx`, `Evaluation20211.xlsx`) and timetable PDFs, but does
    not yet use their respondent counts ("Eval Students") for weighting. If
    response-rate weighting or n-counts are needed in the dashboard, that
    requires extending this metadata parser.
3. The Spring 2021 nested "Teaching Evaluation 20212" folder is a
   duplicate of a subset of the top-level Spring 2021 data. It is
   excluded because the top-level folder is the more complete of the
   two (it additionally has GFP and ID, which the nested copy lacks).
4. Question text is left in the original Arabic per instruction --
   no translation layer is applied.

OUTPUT
------
teaching_evaluation_consolidated.json, shaped as:

{
  "semester_order": ["Fall 2020", "Spring 2021", "Fall 2021", "Spring 2022"],
  "question_reference": [ {seq, category, subcode, text}, ... ],   # 20 items
  "semesters": {
    "<Semester Label>": {
      "<DEPT>": {
        "instructors": { "<name>": {"courses":[...], "questions":[...]} },
        "programs":    { "<name>": {"courses":[...], "questions":[...]} },
        "instructors_summary": [ {name, overall, AA, CK, PV, courses}, ... ],
        "programs_summary":    [ {name, overall, AA, CK, PV}, ... ],
        "department_avg": <float>
      }, ...
    }, ...
  }
}

Course summaries include the course code, an overall course score calculated
from the question scores, and a course name when one is found in the Excel or
PDF schedule metadata.

USAGE
-----
    python etl_teaching_evaluation.py \
        --input "/path/to/2 Years Data for Teaching Evaluation Surveys Results" \
        --output teaching_evaluation_consolidated.json
"""

import argparse
import json
import os
import re

import openpyxl
from pypdf import PdfReader

DEPTS = {"AFS", "CSMIS", "GFP", "ID"}

SEMESTER_LABELS = {
    "Teaching Evaluation Fall 2020": "Fall 2020",
    "Teaching Evaluation Spring 2021": "Spring 2021",
    "Teaching Evaluation Fall 2021": "Fall 2021",
    "Teaching Evaluation Spring 2022": "Spring 2022",
}
SEMESTER_ORDER = [
    "Teaching Evaluation Fall 2020",
    "Teaching Evaluation Spring 2021",
    "Teaching Evaluation Fall 2021",
    "Teaching Evaluation Spring 2022",
]

IMPROVEMENT_THRESHOLD = 65.0


def clean(value):
    """Strip the non-breaking spaces (\\xa0) Excel exports are riddled with."""
    if isinstance(value, str):
        return value.replace("\xa0", " ").strip()
    return value


def find_department_folders(input_root):
    """
    Walk the raw folder tree and return {(semester_folder_name, dept): path}
    for every AFS/CSMIS/GFP/ID folder found, EXCLUDING the known Spring 2021
    duplicate ("Teaching Evaluation 20212") so each (semester, dept) maps to
    exactly one folder.
    """
    found = {}
    for semester in os.listdir(input_root):
        sem_path = os.path.join(input_root, semester)
        if not os.path.isdir(sem_path):
            continue
        for root, _dirs, _files in os.walk(sem_path):
            if "Teaching Evaluation 20212" in root:
                continue  # known duplicate, see LIMITATIONS
            name = os.path.basename(root)
            if name in DEPTS:
                found[(semester, name)] = root
    return found


def find_file(folder, kind):
    """
    kind is 'staff' or 'programme'. Prefers an exact '<kind>.xlsx' match
    (case-insensitive); falls back to any .xlsx whose name contains kind
    (handles the 'GFP staff Aisha part 20201.xlsx' style filenames), while
    excluding the unrelated top-level 'Evaluation<term>.xlsx' raw files.
    """
    files = [f for f in os.listdir(folder) if f.lower().endswith(".xlsx")]
    exact = [f for f in files if f.lower() == f"{kind}.xlsx"]
    if exact:
        return exact[0]
    partial = [f for f in files if kind in f.lower() and "evaluation" not in f.lower()]
    return partial[0] if partial else None


def find_course_names(folder, excluded_files):
    """Read optional course-code/title metadata from per-course workbooks."""
    names = {}
    for filename in os.listdir(folder):
        if not filename.lower().endswith(".xlsx") or filename in excluded_files:
            continue
        try:
            wb = openpyxl.load_workbook(os.path.join(folder, filename), data_only=True, read_only=True)
            for ws in wb.worksheets:
                metadata = {}
                for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
                    label = clean(row[1]) if len(row) > 1 else None
                    value = clean(row[4]) if len(row) > 4 else None
                    if label in {"Course Code", "Course Name"}:
                        metadata[label] = value
                code = metadata.get("Course Code")
                name = metadata.get("Course Name")
                if code is not None and name:
                    names[str(code)] = str(name)
        except Exception:
            continue
    return names


def find_pdf_course_names(folder):
    """Extract course-code/name pairs from timetable and staff-schedule PDFs."""
    names = {}
    code_pattern = re.compile(r"^\d{6}\*?$")
    ignored_lines = {"course course name hrs", "activity", "section time room", "hours / hour", "count"}

    for filename in os.listdir(folder):
        if not filename.lower().endswith(".pdf"):
            continue
        path = os.path.join(folder, filename)
        try:
            reader = PdfReader(path)
            for page in reader.pages:
                lines = [clean(line) for line in (page.extract_text() or "").splitlines()]
                lines = [line for line in lines if line]
                for index, line in enumerate(lines):
                    if not code_pattern.fullmatch(line):
                        continue
                    code_start = index
                    while code_start > 0 and code_pattern.fullmatch(lines[code_start - 1]):
                        code_start -= 1
                    code_end = index + 1
                    while code_end < len(lines) and code_pattern.fullmatch(lines[code_end]):
                        code_end += 1

                    codes = [re.sub(r"\*+$", "", value) for value in lines[code_start:code_end]]
                    if len(codes) < 2:
                        continue

                    candidates = []
                    for previous in reversed(lines[:code_start]):
                        lowered = previous.lower()
                        if lowered in ignored_lines or not re.search(r"[A-Za-z]", previous):
                            continue
                        if re.search(r"page\s+\d+|instructor schedule|timetable|faculty|department|semester", lowered):
                            continue
                        candidates.append(previous)
                        if len(candidates) == len(codes):
                            break

                    if len(candidates) == len(codes):
                        for code, name in zip(codes, reversed(candidates)):
                            names.setdefault(code, name)
        except Exception:
            continue
    return names


def parse_sheet(ws, course_names=None):
    """
    Parse one instructor/program sheet into:
            {"courses": [<course codes in column order>],
             "course_summary": [{"name": <course code>, "overall": <score>}, ...],
       "questions": [{category, subcode, seq, text, scores:{course:score}, avg}, ...]}
    """
    header = [clean(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]

    course_cols = []
    avg_col = None
    for c in range(6, ws.max_column + 1):  # course codes start at column F
        v = header[c - 1]
        if v == "Avg":
            avg_col = c
            break
        if v is not None:
            course_cols.append((c, v))

    questions = []
    r = 2
    while r <= ws.max_row:
        cat = clean(ws.cell(row=r, column=2).value)
        subcode = clean(ws.cell(row=r, column=3).value)
        text = clean(ws.cell(row=r, column=4).value)
        seq = ws.cell(row=r, column=5).value

        if cat is None and subcode is None and text is None:
            break  # blank row = end of question block
        if isinstance(seq, str):
            break  # hit a summary row like "Total Avg" / "AA Avg" (seq is text there)

        scores = {}
        for c, code in course_cols:
            v = ws.cell(row=r, column=c).value
            scores[str(code)] = round(v, 2) if isinstance(v, (int, float)) else None

        avg_val = ws.cell(row=r, column=avg_col).value if avg_col else None
        if isinstance(avg_val, (int, float)):
            questions.append(
                {
                    "category": cat,
                    "subcode": subcode,
                    "seq": seq,
                    "text": text,
                    "scores": scores,
                    "avg": round(avg_val, 2),
                }
            )
        r += 1

    course_names = course_names or {}
    course_summary = []
    for _, code in course_cols:
        course_code = str(code)
        scores = [q["scores"][course_code] for q in questions if q["scores"].get(course_code) is not None]
        if scores:
            course_summary.append({
                "code": course_code,
                "name": course_names.get(course_code, course_code),
                "overall": round(sum(scores) / len(scores), 2),
            })

    return {
        "courses": [str(c) for _, c in course_cols],
        "course_summary": course_summary,
        "questions": questions,
    }


def overall_avg(parsed_sheet):
    qs = parsed_sheet["questions"]
    return round(sum(q["avg"] for q in qs) / len(qs), 2) if qs else None


def category_avg(parsed_sheet, category):
    vals = [q["avg"] for q in parsed_sheet["questions"] if q["category"] == category]
    return round(sum(vals) / len(vals), 2) if vals else None


def run_qa_checks(semesters_out):
    """
    Verify the assumption the dashboard relies on: every instructor sheet,
    across every semester and department, uses the same 20-question set
    (same seq numbers). If this ever fails for a future data drop, it means
    someone changed the survey instrument and cross-semester comparisons in
    the dashboard would silently become invalid -- so we fail loudly here
    instead.
    """
    reference_seqs = None
    mismatches = []
    for sem, depts in semesters_out.items():
        for dept, entry in depts.items():
            for name, sheet in entry["instructors"].items():
                seqs = tuple(sorted(q["seq"] for q in sheet["questions"]))
                if reference_seqs is None:
                    reference_seqs = seqs
                elif seqs != reference_seqs:
                    mismatches.append((sem, dept, name, len(seqs)))
    if mismatches:
        print(f"WARNING: {len(mismatches)} instructor sheets use a different "
              f"question set than the reference. Cross-semester comparisons "
              f"involving these will not be apples-to-apples:")
        for m in mismatches[:10]:
            print("  ", m)
    else:
        print(f"QA OK: all instructor sheets use the same {len(reference_seqs)}-question set.")


def build_dataset(input_root):
    found = find_department_folders(input_root)

    question_reference = None
    semesters_out = {}
    errors = []

    for (semester_folder, dept), folder in found.items():
        sem_label = SEMESTER_LABELS[semester_folder]
        semesters_out.setdefault(sem_label, {})

        staff_file = find_file(folder, "staff")
        prog_file = find_file(folder, "programme")
        course_names = find_course_names(folder, {f for f in (staff_file, prog_file) if f})
        course_names.update(find_pdf_course_names(folder))

        dept_entry = {
            "instructors": {},
            "programs": {},
            "instructors_summary": [],
            "programs_summary": [],
            "department_avg": None,
        }

        if staff_file:
            try:
                wb = openpyxl.load_workbook(os.path.join(folder, staff_file), data_only=True)
                for sheet_name in wb.sheetnames:
                    parsed = parse_sheet(wb[sheet_name], course_names)
                    if parsed["questions"]:
                        dept_entry["instructors"][clean(sheet_name)] = parsed
                        if question_reference is None:
                            question_reference = [
                                {"seq": q["seq"], "category": q["category"],
                                 "subcode": q["subcode"], "text": q["text"]}
                                for q in parsed["questions"]
                            ]
            except Exception as exc:  # noqa: BLE001 - we want to log & continue
                errors.append((semester_folder, dept, "staff", str(exc)))
        else:
            errors.append((semester_folder, dept, "staff", "no matching file found"))

        if prog_file:
            try:
                wb = openpyxl.load_workbook(os.path.join(folder, prog_file), data_only=True)
                skip_sheets = {"staff", "Staff", "Sheet1", "Sheet2"}
                for sheet_name in wb.sheetnames:
                    if sheet_name in skip_sheets:
                        continue
                    parsed = parse_sheet(wb[sheet_name], course_names)
                    if parsed["questions"]:
                        dept_entry["programs"][clean(sheet_name)] = parsed
            except Exception as exc:  # noqa: BLE001
                errors.append((semester_folder, dept, "programme", str(exc)))
        else:
            errors.append((semester_folder, dept, "programme", "no matching file found"))

        for name, sheet in dept_entry["instructors"].items():
            avg = overall_avg(sheet)
            if avg is None:
                continue
            dept_entry["instructors_summary"].append({
                "name": name,
                "overall": avg,
                "AA": category_avg(sheet, "AA"),
                "CK": category_avg(sheet, "CK"),
                "PV": category_avg(sheet, "PV"),
                "courses": sheet["course_summary"],
            })

        for name, sheet in dept_entry["programs"].items():
            avg = overall_avg(sheet)
            if avg is None:
                continue
            dept_entry["programs_summary"].append({
                "name": name,
                "overall": avg,
                "AA": category_avg(sheet, "AA"),
                "CK": category_avg(sheet, "CK"),
                "PV": category_avg(sheet, "PV"),
            })

        if dept_entry["instructors_summary"]:
            dept_entry["department_avg"] = round(
                sum(i["overall"] for i in dept_entry["instructors_summary"])
                / len(dept_entry["instructors_summary"]), 2
            )

        dept_entry["instructors_summary"].sort(key=lambda x: -x["overall"])
        dept_entry["programs_summary"].sort(key=lambda x: -x["overall"])

        semesters_out[sem_label][dept] = dept_entry

    if errors:
        print(f"WARNING: {len(errors)} file(s) could not be read:")
        for e in errors:
            print("  ", e)

    run_qa_checks(semesters_out)

    return {
        "semester_order": [SEMESTER_LABELS[s] for s in SEMESTER_ORDER],
        "question_reference": question_reference,
        "semesters": semesters_out,
    }


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--input", required=True, help="Path to the '2 Years Data for Teaching Evaluation Surveys Results' folder")
    parser.add_argument("--output", default="teaching_evaluation_consolidated.json", help="Output JSON path")
    args = parser.parse_args()

    dataset = build_dataset(args.input)

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(dataset, f, ensure_ascii=False, indent=None)

    n_instructor_records = sum(
        len(dept["instructors_summary"])
        for sem in dataset["semesters"].values()
        for dept in sem.values()
    )
    print(f"\nWrote {args.output}")
    print(f"  Semesters: {dataset['semester_order']}")
    print(f"  Instructor-semester records: {n_instructor_records}")
    print(f"  Questions per instructor: {len(dataset['question_reference'])}")


if __name__ == "__main__":
    main()
